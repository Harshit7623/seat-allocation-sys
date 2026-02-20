import json
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import Dict, List, Any
import os


class JSONToExcelConverter:
    """
    A converter that transforms JSON seating plan data to Excel format.
    """
    
    def __init__(self):
        self.column_widths = {
            'room_no': 10,
            'position': 8,
            'batch_label': 12,
            'batch': 6,
            'paper_set': 8,
            'roll_number': 16,
            'student_name': 30,
            'display': 18,
            'block': 6,
            'is_broken': 10,
            'is_unallocated': 12,
            'color': 12,
            'css_class': 20,
            'degree': 10,
            'branch': 8,
            'joining_year': 12
        }
        
        # Define colors for highlighting
        self.highlight_colors = {
            'broken_seat': 'FF9999',  # Light red
            'unallocated': 'FFCC99',  # Light orange
            'batch_header': 'E6E6FA',  # Lavender
            'room_header': 'D8E8F8',  # Light blue
        }
        
    def load_json(self, json_file_path: str) -> Dict:
        """
        Load JSON data from file.
        """
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            print(f"Successfully loaded JSON from {json_file_path}")
            print(f"Total students: {data.get('metadata', {}).get('total_students', 'N/A')}")
            print(f"Active rooms: {data.get('metadata', {}).get('active_rooms', [])}")
            return data
        except Exception as e:
            print(f"Error loading JSON file: {e}")
            raise
    
    def load_json_string(self, json_string: str) -> Dict:
        """
        Load JSON data from string.
        """
        try:
            data = json.loads(json_string)
            print("Successfully loaded JSON from string")
            print(f"Total students: {data.get('metadata', {}).get('total_students', 'N/A')}")
            return data
        except Exception as e:
            print(f"Error parsing JSON string: {e}")
            raise
    
    def extract_all_students(self, data: Dict) -> List[Dict]:
        """
        Extract all students from all rooms into a flat list.
        """
        all_students = []
        rooms_data = data.get('rooms', {})
        
        for room_name, room_info in rooms_data.items():
            batches = room_info.get('batches', {})
            
            for batch_name, batch_info in batches.items():
                students = batch_info.get('students', [])
                
                for student in students:
                    # Add batch info to each student
                    student_copy = student.copy()
                    student_copy['room_name'] = room_name
                    student_copy['batch_info_degree'] = batch_info.get('info', {}).get('degree', '')
                    student_copy['batch_info_branch'] = batch_info.get('info', {}).get('branch', '')
                    student_copy['batch_info_joining_year'] = batch_info.get('info', {}).get('joining_year', '')
                    
                    all_students.append(student_copy)
        
        print(f"Extracted {len(all_students)} students from {len(rooms_data)} rooms")
        return all_students
    
    def create_student_dataframe(self, students: List[Dict]) -> pd.DataFrame:
        """
        Create a pandas DataFrame from student data.
        """
        # Define column order
        columns = [
            'room_name', 'position', 'batch_label', 'batch', 
            'paper_set', 'roll_number', 'student_name', 'display',
            'block', 'is_broken', 'is_unallocated', 'color',
            'css_class', 'batch_info_degree', 'batch_info_branch',
            'batch_info_joining_year'
        ]
        
        # Create DataFrame
        df = pd.DataFrame(students)
        
        # Ensure all columns exist (add missing ones)
        for col in columns:
            if col not in df.columns:
                df[col] = ''
        
        # Reorder columns
        df = df[columns]
        
        return df
    
    def create_batch_summary(self, data: Dict) -> pd.DataFrame:
        """
        Create a summary of batches across all rooms.
        """
        summary_data = []
        rooms_data = data.get('rooms', {})
        
        for room_name, room_info in rooms_data.items():
            batches = room_info.get('batches', {})
            student_count = room_info.get('student_count', 0)
            
            for batch_name, batch_info in batches.items():
                info = batch_info.get('info', {})
                students = batch_info.get('students', [])
                
                summary_data.append({
                    'room_name': room_name,
                    'batch_label': batch_name,
                    'degree': info.get('degree', ''),
                    'branch': info.get('branch', ''),
                    'joining_year': info.get('joining_year', ''),
                    'student_count': len(students),
                    'paper_set_a': len([s for s in students if s.get('paper_set') == 'A']),
                    'paper_set_b': len([s for s in students if s.get('paper_set') == 'B']),
                    'block_distribution': ', '.join([str(s.get('block', '')) for s in students[:5]]) + ('...' if len(students) > 5 else '')
                })
        
        # Add total row
        if summary_data:
            total_students = sum(item['student_count'] for item in summary_data)
            summary_data.append({
                'room_name': 'TOTAL',
                'batch_label': '',
                'degree': '',
                'branch': '',
                'joining_year': '',
                'student_count': total_students,
                'paper_set_a': sum(item['paper_set_a'] for item in summary_data),
                'paper_set_b': sum(item['paper_set_b'] for item in summary_data),
                'block_distribution': ''
            })
        
        df = pd.DataFrame(summary_data)
        return df
    
    def create_room_summary(self, data: Dict) -> pd.DataFrame:
        """
        Create a summary of rooms.
        """
        summary_data = []
        rooms_data = data.get('rooms', {})
        
        for room_name, room_info in rooms_data.items():
            batches = room_info.get('batches', {})
            student_count = room_info.get('student_count', 0)
            optimized = room_info.get('optimized', False)
            inputs = room_info.get('inputs', {})
            
            batch_list = list(batches.keys())
            
            summary_data.append({
                'room_name': room_name,
                'student_count': student_count,
                'batch_count': len(batches),
                'batches': ', '.join(batch_list),
                'optimized': 'Yes' if optimized else 'No',
                'rows': inputs.get('rows', ''),
                'cols': inputs.get('cols', ''),
                'block_width': inputs.get('block_width', ''),
                'broken_seats_count': len(inputs.get('broken_seats', []))
            })
        
        # Add total row
        if summary_data:
            total_students = sum(item['student_count'] for item in summary_data)
            summary_data.append({
                'room_name': 'TOTAL',
                'student_count': total_students,
                'batch_count': sum(item['batch_count'] for item in summary_data),
                'batches': '',
                'optimized': '',
                'rows': '',
                'cols': '',
                'block_width': '',
                'broken_seats_count': sum(item['broken_seats_count'] for item in summary_data)
            })
        
        df = pd.DataFrame(summary_data)
        return df
    
    def create_seating_matrix_sheets(self, data: Dict, workbook) -> None:
        """
        Create separate sheets for each room's seating matrix.
        """
        rooms_data = data.get('rooms', {})
        
        for room_name, room_info in rooms_data.items():
            raw_matrix = room_info.get('raw_matrix', [])
            if not raw_matrix:
                continue
            
            # Limit sheet name to 31 characters (Excel limit)
            sheet_name = f"Matrix_{room_name}"[:31]
            
            # Skip if sheet already exists
            if sheet_name in workbook.sheetnames:
                sheet_name = f"{sheet_name}_{len([s for s in workbook.sheetnames if s.startswith(sheet_name)])}"
            
            ws = workbook.create_sheet(title=sheet_name)
            
            # Write headers
            headers = ['Row', 'Position', 'Batch', 'Roll Number', 'Student Name', 'Paper Set', 'Block', 'Status']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.highlight_colors['room_header'],
                                       end_color=self.highlight_colors['room_header'],
                                       fill_type='solid')
            
            row_idx = 2
            for row_num, row_data in enumerate(raw_matrix, 1):
                for seat_data in row_data:
                    ws.cell(row=row_idx, column=1, value=row_num)
                    ws.cell(row=row_idx, column=2, value=seat_data.get('position', ''))
                    ws.cell(row=row_idx, column=3, value=seat_data.get('batch_label', ''))
                    ws.cell(row=row_idx, column=4, value=seat_data.get('roll_number', ''))
                    ws.cell(row=row_idx, column=5, value=seat_data.get('student_name', ''))
                    ws.cell(row=row_idx, column=6, value=seat_data.get('paper_set', ''))
                    ws.cell(row=row_idx, column=7, value=seat_data.get('block', ''))
                    
                    # Status column
                    status = []
                    if seat_data.get('is_broken', False):
                        status.append('Broken Seat')
                    if seat_data.get('is_unallocated', False):
                        status.append('Unallocated')
                    ws.cell(row=row_idx, column=8, value=', '.join(status) if status else 'OK')
                    
                    # Apply formatting for broken/unallocated seats
                    if seat_data.get('is_broken', False):
                        for col in range(1, 9):
                            ws.cell(row=row_idx, column=col).fill = PatternFill(
                                start_color=self.highlight_colors['broken_seat'],
                                end_color=self.highlight_colors['broken_seat'],
                                fill_type='solid'
                            )
                    elif seat_data.get('is_unallocated', False):
                        for col in range(1, 9):
                            ws.cell(row=row_idx, column=col).fill = PatternFill(
                                start_color=self.highlight_colors['unallocated'],
                                end_color=self.highlight_colors['unallocated'],
                                fill_type='solid'
                            )
                    
                    row_idx += 1
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def apply_styling(self, ws, df: pd.DataFrame) -> None:
        """
        Apply styling to the worksheet.
        """
        # Apply column widths
        for i, column in enumerate(df.columns, 1):
            col_letter = openpyxl.utils.get_column_letter(i)
            if column in self.column_widths:
                ws.column_dimensions[col_letter].width = self.column_widths[column]
            else:
                ws.column_dimensions[col_letter].width = 15
        
        # Style header row
        for i in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply alternating row colors and highlight broken/unallocated seats
        for row in range(2, ws.max_row + 1):
            is_broken = ws.cell(row=row, column=df.columns.get_loc('is_broken') + 1).value
            is_unallocated = ws.cell(row=row, column=df.columns.get_loc('is_unallocated') + 1).value
            
            if is_broken == True or (isinstance(is_broken, str) and is_broken.lower() == 'true'):
                fill_color = self.highlight_colors['broken_seat']
            elif is_unallocated == True or (isinstance(is_unallocated, str) and is_unallocated.lower() == 'true'):
                fill_color = self.highlight_colors['unallocated']
            elif row % 2 == 0:
                fill_color = "F2F2F2"  # Light gray for even rows
            else:
                fill_color = "FFFFFF"  # White for odd rows
            
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border
    
    def convert_to_excel(self, json_data: Dict, output_path: str) -> str:
        """
        Main conversion method: JSON to Excel.
        """
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 1. Extract all students
                all_students = self.extract_all_students(json_data)
                student_df = self.create_student_dataframe(all_students)
                
                # 2. Create batch summary
                batch_summary_df = self.create_batch_summary(json_data)
                
                # 3. Create room summary
                room_summary_df = self.create_room_summary(json_data)
                
                # 4. Write DataFrames to Excel
                student_df.to_excel(writer, sheet_name='All_Students', index=False)
                batch_summary_df.to_excel(writer, sheet_name='Batch_Summary', index=False)
                room_summary_df.to_excel(writer, sheet_name='Room_Summary', index=False)
                
                # 5. Create metadata sheet
                metadata = json_data.get('metadata', {})
                inputs = json_data.get('inputs', {})
                
                metadata_data = {
                    'Field': list(metadata.keys()) + [''] + list(inputs.keys()),
                    'Value': list(metadata.values()) + [''] + list(inputs.values())
                }
                metadata_df = pd.DataFrame(metadata_data)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)
                
                # Save workbook to apply additional formatting
                writer._save()
            
            # Reopen workbook for additional formatting and seating matrices
            workbook = openpyxl.load_workbook(output_path)
            
            # Apply styling to each sheet
            for sheet_name in ['All_Students', 'Batch_Summary', 'Room_Summary', 'Metadata']:
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    if sheet_name == 'All_Students':
                        self.apply_styling(ws, student_df)
                    elif sheet_name == 'Batch_Summary':
                        self.apply_styling(ws, batch_summary_df)
                    elif sheet_name == 'Room_Summary':
                        self.apply_styling(ws, room_summary_df)
                    else:
                        # Simple formatting for metadata
                        for i in range(1, 3):
                            col_letter = openpyxl.utils.get_column_letter(i)
                            ws.column_dimensions[col_letter].width = 25
            
            # Create seating matrix sheets
            self.create_seating_matrix_sheets(json_data, workbook)
            
            # Save final workbook
            workbook.save(output_path)
            
            print(f"Excel file created successfully: {output_path}")
            print(f"Sheets created: {workbook.sheetnames}")
            
            return output_path
            
        except Exception as e:
            print(f"Error converting to Excel: {e}")
            raise
    
    def convert_json_file(self, json_file_path: str, output_excel_path: str = None) -> str:
        """
        Convert JSON file to Excel.
        """
        # Load JSON data
        data = self.load_json(json_file_path)
        
        # Generate output path if not provided
        if output_excel_path is None:
            base_name = os.path.splitext(os.path.basename(json_file_path))[0]
            output_excel_path = f"{base_name}_seating_plan.xlsx"
        
        # Convert to Excel
        return self.convert_to_excel(data, output_excel_path)
    
    def convert_json_string(self, json_string: str, output_excel_path: str = "seating_plan.xlsx") -> str:
        """
        Convert JSON string to Excel.
        """
        # Load JSON data
        data = self.load_json_string(json_string)
        
        # Convert to Excel
        return self.convert_to_excel(data, output_excel_path)


# Example usage function
def example_usage():
    """
    Example of how to use the JSONToExcelConverter class.
    """
    converter = JSONToExcelConverter()
    
    # Option 1: Convert from JSON file
    # converter.convert_json_file('seating_plan.json', 'output.xlsx')
    
    # Option 2: Convert from JSON string (like in your example)
    # with open('seating_plan.json', 'r') as f:
    #     json_string = f.read()
    # converter.convert_json_string(json_string, 'output.xlsx')
    
    print("JSONToExcelConverter is ready to use!")
    print("\nUsage:")
    print("1. Create converter: converter = JSONToExcelConverter()")
    print("2. Convert from file: converter.convert_json_file('input.json', 'output.xlsx')")
    print("3. Convert from string: converter.convert_json_string(json_str, 'output.xlsx')")


if __name__ == "__main__":
    # Run example usage when module is executed directly
    example_usage()